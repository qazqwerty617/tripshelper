import openpyxl
import sys
from config import EXCEL_PATH

# Set stdout to utf-8 for windows terminal
sys.stdout.reconfigure(encoding='utf-8')

def dump_tax_sheet():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    for sheet_name in wb.sheetnames:
        if "ПОДАТОК" in sheet_name.upper():
            ws = wb[sheet_name]
            print(f"--- Sheet: {sheet_name} ---")
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), 1):
                if any(row):
                    print(f"{i}: {row}")

if __name__ == "__main__":
    dump_tax_sheet()
