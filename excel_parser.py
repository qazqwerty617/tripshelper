import openpyxl
import os
import logging
import re
from typing import List, Dict
from config import EXCEL_PATH
import time

logger = logging.getLogger(__name__)

# Cache for hotel database and tax data
_db_cache = {
    "data": None,
    "last_loaded": 0,
    "file_mtime": 0
}

def get_hotel_db() -> Dict[str, List[Dict[str, str]]]:
    global _db_cache
    
    if not os.path.exists(EXCEL_PATH):
        logger.warning(f"Excel file not found at {EXCEL_PATH}")
        return {}

    # Check if we can use cache
    current_mtime = os.path.getmtime(EXCEL_PATH)
    if (_db_cache["data"] is not None and 
        _db_cache["file_mtime"] == current_mtime and 
        time.time() - _db_cache["last_loaded"] < 3600): # 1 hour cache
        return _db_cache["data"]

    try:
        # read_only=True DOES NOT support hyperlinks. We must use read_only=False.
        # data_only=False to see formulas like =HYPERLINK(...)
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False, read_only=False)
        db = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            normalized_sheet = sheet_name.strip().lower()
            hotels = []

            # Use ws.rows or iter_rows without values_only to get Cell objects
            for row in ws.iter_rows():
                if not any(cell.value for cell in row if cell.value is not None):
                    continue
                
                hotel_name = ""
                link = ""

                for cell in row:
                    val = cell.value
                    if val is None:
                        continue
                    
                    s_val = str(val).strip()
                    if not s_val: continue

                    # 1. Try to get link from cell.hyperlink (the blue underlined text)
                    current_link = ""
                    if cell.hyperlink and cell.hyperlink.target:
                        target = str(cell.hyperlink.target).strip()
                        if target.startswith("http") or target.startswith("www"):
                            current_link = target

                    # 2. Try to extract from =HYPERLINK formula
                    if not current_link and s_val.startswith("=HYPERLINK"):
                        match = re.search(r'=HYPERLINK\("([^"]+)"', s_val, re.IGNORECASE)
                        if match:
                            current_link = match.group(1)
                            # If we don't have a name yet, try to get the "friendly name"
                            if not hotel_name:
                                name_match = re.search(r', ?"([^"]+)"\)', s_val)
                                if name_match:
                                    hotel_name = name_match.group(1)

                    # 3. If it's just a plain text link
                    if not current_link and (s_val.startswith("http") or s_val.startswith("www")):
                        current_link = s_val

                    # Update row link if we found one
                    if current_link:
                        link = current_link
                        continue # Move to next cell, don't treat link cell as a name

                    # 4. If it's a potential hotel name (not a digit, not a formula, not a link)
                    # and we don't have a name yet for this row
                    if not hotel_name:
                        if not s_val.isdigit() and not s_val.startswith("="):
                            if len(s_val) > 2:
                                hotel_name = s_val

                if hotel_name:
                    hotels.append({"hotel": hotel_name, "link": link or "Посилання відсутнє ⚠️"})

            if hotels and "податок" not in normalized_sheet:
                db[normalized_sheet] = hotels

        # Update cache
        _db_cache["data"] = db
        _db_cache["file_mtime"] = current_mtime
        _db_cache["last_loaded"] = time.time()
        
        return db

    except Exception as e:
        logger.error(f"Error reading excel: {e}")
        return {}


def format_hotel_db_for_prompt() -> str:
    db = get_hotel_db()
    if not db:
        return "База готелів порожня або файл не знайдено."

    lines = []
    for sheet, hotels in db.items():
        for h in hotels:
            lines.append(f"Назва: {h['hotel']} | Посилання: {h['link']}")
    return "\n".join(lines)


def get_tourist_tax_db() -> str:
    if not os.path.exists(EXCEL_PATH):
        return ""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=False)
        tax_md = []
        for sheet_name in wb.sheetnames:
            if "ПОДАТОК" in sheet_name.upper():
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
                    if not any(row): continue

                    # Check if it's a title row or notes row
                    if row[2] and not row[3] and not row[5]:
                        tax_md.append(f"**{str(row[2]).strip()}**")
                        continue

                    # Table header or data row — data at [3..11]
                    if row[3] and row[4] and row[5] is not None:
                        resort = str(row[3]).strip()
                        period = str(row[4]).strip()
                        unit = str(row[5]).strip() if row[5] else ""
                        stars_0 = str(row[6]).strip() if row[6] is not None else ""
                        stars_1_2 = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
                        stars_3 = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
                        stars_4 = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ""
                        stars_5 = str(row[10]).strip() if len(row) > 10 and row[10] is not None else ""
                        who_pays = str(row[11]).strip() if len(row) > 11 and row[11] else ""

                        if resort.upper() == 'КУРОРТ':
                            tax_md.append(f"| {resort} | {period} | {unit} | БЕЗ ЗІРОК | 1-2★ | 3★ | 4★ | 5★ | {who_pays} |")
                            tax_md.append(f"|---|---|---|---|---|---|---|---|---|")
                        else:
                            tax_md.append(f"| {resort} | {period} | {unit} | {stars_0} | {stars_1_2} | {stars_3} | {stars_4} | {stars_5} | {who_pays} |")
                break

        if tax_md:
            return ("ІНФОРМАЦІЯ ПРО ТУРИСТИЧНИЙ ПОДАТОК:\n" + "\n".join(tax_md) +
                    "\n\n❌ КРИТИЧНО ВАЖЛИВО: Якщо напрямку НЕМАЄ в таблицях вище (наприклад: Туреччина, Анталія, Єгипет, ОАЕ, Кіпр, Сонячний Берег, Золоті Піски, Тенеріфе, Гран-Канарія, Фуертевентура, Лансароте, Коста-дель-Соль) — туристичного податку НЕ ІСНУЄ! НЕ ВИГАДУЙ податок! Встанови його рівним 0€.")
        return ""
    except Exception as e:
        logger.error(f"Error reading tourist tax: {e}")
        return ""


# Maps destination keywords to resort name in the tax table
_RESORT_ALIASES = {
    "майорк": "Майорка", "mallorca": "Майорка",
    "ібіц": "Ібіца", "ibiza": "Ібіца",
    "коста-брав": "Коста-Брава",
    "коста-дорад": "Коста-Дорада",
    "мальт": "Мальта", "malta": "Мальта",
    "рімін": "Ріміні", "rimini": "Ріміні",
    "лідо": "Лідо-ді-Єзоло", "jesolo": "Лідо-ді-Єзоло",
    "мадейр": "Мадейра", "madeira": "Мадейра",
    "крит": "Крит", "crete": "Крит",
    "корфу": "Корфу", "corfu": "Корфу",
    "родос": "Родос", "rhodes": "Родос",
    "міконос": "Міконос", "mykonos": "Міконос",
    "халкідік": "Халкідікі", "halkidiki": "Халкідікі",
    "закінтос": "Закінтос", "zakynthos": "Закінтос",
    "тенериф": "Тенеріфе", "tenerife": "Тенеріфе",
}


def _month_in_period(period: str, month: int) -> bool:
    """Check if month (1-12) falls within a period string like '01.05–31.10' or 'цілий рік'."""
    period = period.strip().lower()
    if "цілий" in period:
        return True
    m = re.search(r'(\d{2})\.(\d{2})[–\-](\d{2})\.(\d{2})', period)
    if m:
        s, e = int(m.group(2)), int(m.group(4))
        return (s <= month <= e) if s <= e else (month >= s or month <= e)
    m2 = re.search(r'з\s+\d{2}\.(\d{2})', period)
    if m2:
        return month >= int(m2.group(1))
    m3 = re.search(r'до\s+\d{2}\.(\d{2})', period)
    if m3:
        return month <= int(m3.group(1))
    return False


def get_tax_per_person_per_night(destination: str, stars: int, month: int, num_people: int = 1) -> float:
    """
    Return tourist tax per person per night for given destination/stars/month.
    Columns in Excel: [3]=resort [4]=period [5]=unit [6]=0★ [7]=1-2★ [8]=3★ [9]=4★ [10]=5★
    Returns 0.0 if destination has no tax.
    """
    if not os.path.exists(EXCEL_PATH):
        return 0.0
    try:
        dest_lower = destination.lower()
        resort_name = None
        for key, val in _RESORT_ALIASES.items():
            if key in dest_lower:
                resort_name = val.lower()
                break
        
        # If no alias, use the destination itself as a fallback
        if not resort_name:
            resort_name = destination.strip().lower()

        # Use read_only=False to be consistent with get_hotel_db and ensure access if needed
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=False)
        for sheet_name in wb.sheetnames:
            if "ПОДАТОК" not in sheet_name.upper():
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=100, values_only=True): # Increased max_row
                if not any(row):
                    continue
                if len(row) < 11 or not row[3] or not row[4]:
                    continue
                
                row_resort = str(row[3]).strip().lower()
                if row_resort in ("курорт", "таблиця", "ставки", "курорти", "назва"):
                    continue
                
                # logger.debug(f"Comparing resort: '{resort_name}' with '{row_resort}'")
                
                # Flexible match: either exact or one contains the other
                if resort_name not in row_resort and row_resort not in resort_name:
                    continue
                
                if not _month_in_period(str(row[4]), month):
                    logger.info(f"Tax period mismatch: '{row[4]}' for month {month}")
                    continue
                
                unit = str(row[5]).strip().lower() if row[5] else ""
                # Col mapping: 0★→6, 1-2★→7, 3★→8, 4★→9, 5★→10
                # UPDATE: 1-2 stars now count as "no stars" (0★) as requested
                col = {0: 6, 1: 6, 2: 6, 3: 8, 4: 9, 5: 10}.get(stars, 9)
                rate_val = row[col] if len(row) > col else None
                
                try:
                    # Clean the rate value (sometimes it might have currency symbols or commas)
                    if isinstance(rate_val, str):
                        rate_val = rate_val.replace(',', '.').replace('€', '').strip()
                    rate = float(rate_val) if rate_val is not None and str(rate_val).strip() != "" else 0.0
                except (ValueError, TypeError):
                    rate = 0.0
                
                # If "за номер / ніч" — divide by people to get per-person rate
                if "номер" in unit and num_people > 0:
                    rate = rate / num_people
                
                logger.info(f"TAX FOUND: {row_resort} {stars}★ month={month} -> {rate}€/person/night")
                return rate
        
        logger.info(f"TAX NOT FOUND for {destination} (resort={resort_name}), stars={stars}, month={month}")
        return 0.0
    except Exception as e:
        logger.error(f"Tax lookup error: {e}")
        return 0.0
